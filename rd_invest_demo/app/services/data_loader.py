from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..config import SOURCE_FILES
from ..db import db_session, reset_business_tables

PROJECT_CODE_PATTERN = re.compile(r"(R[0-9A-Za-z]+)")


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def extract_project_code(text: str) -> str | None:
    matched = PROJECT_CODE_PATTERN.search(text or "")
    return matched.group(1) if matched else None


def normalize_cost_class(raw: str) -> str:
    text = (raw or "").strip().lower()
    if "capex" in text:
        return "CAPEX"
    if "opex" in text:
        return "OPEX"
    return "TOTAL"


def normalize_stage_flag(value: Any) -> int:
    if value in (1, "1", True, "TRUE", "true"):
        return 1
    return 0


def month_from_chinese(text: str, fallback: datetime | None = None) -> str:
    mapping = {
        "一月": "01",
        "二月": "02",
        "三月": "03",
        "四月": "04",
        "五月": "05",
        "六月": "06",
        "七月": "07",
        "八月": "08",
        "九月": "09",
        "十月": "10",
        "十一月": "11",
        "十二月": "12",
    }
    if text in mapping:
        return f"2026-{mapping[text]}"
    if fallback:
        return fallback.strftime("%Y-%m")
    return "2026-01"


def month_from_filename(path: Path) -> str:
    name = path.name
    if "10月31日" in name:
        return "2025-10"
    if "10月XX日" in name:
        return "2025-09"
    return "2025-10"


def ensure_department(conn, dept_name: str) -> str:
    clean_name = dept_name.strip()
    if not clean_name:
        clean_name = "未分配部门"
    row = conn.execute(
        "SELECT dept_code FROM departments WHERE dept_name = ?",
        (clean_name,),
    ).fetchone()
    if row:
        return row["dept_code"]
    existing_count = conn.execute("SELECT COUNT(*) AS c FROM departments").fetchone()["c"]
    dept_code = f"D{existing_count + 1:03d}"
    conn.execute(
        "INSERT INTO departments (dept_code, dept_name) VALUES (?, ?)",
        (dept_code, clean_name),
    )
    return dept_code


def upsert_project(
    conn,
    project_code: str,
    project_name: str,
    dept_code: str | None,
    category: str | None,
    plan_year: int | None,
    source_file: str,
) -> None:
    conn.execute(
        """
        INSERT INTO projects (project_code, project_name, dept_code, category, plan_year, source_file)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(project_code) DO UPDATE SET
            project_name = excluded.project_name,
            dept_code = COALESCE(excluded.dept_code, projects.dept_code),
            category = COALESCE(excluded.category, projects.category),
            plan_year = COALESCE(excluded.plan_year, projects.plan_year),
            source_file = excluded.source_file
        """,
        (project_code, project_name, dept_code, category, plan_year, source_file),
    )


def insert_import_log(conn, source_name: str, loaded_rows: int, note: str = "") -> None:
    conn.execute(
        """
        INSERT INTO import_log (source_name, loaded_rows, loaded_at, note)
        VALUES (?, ?, ?, ?)
        """,
        (source_name, loaded_rows, datetime.now().isoformat(timespec="seconds"), note),
    )


def load_dept_target(conn) -> int:
    path = SOURCE_FILES["dept_target"]
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    loaded = 0
    company_total = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        dept_name = normalize_text(row[1])
        target_hkd = to_float(row[2])
        if not dept_name or target_hkd is None:
            continue
        dept_code = ensure_department(conn, dept_name)
        conn.execute(
            """
            INSERT INTO kpi_targets (year, scope_type, scope_key, cost_class, target_hkd, source_file)
            VALUES (?, 'department', ?, 'TOTAL', ?, ?)
            """,
            (2026, dept_code, target_hkd * 10000, path.name),
        )
        loaded += 1
        company_total += target_hkd * 10000
    if company_total > 0:
        conn.execute(
            """
            INSERT INTO kpi_targets (year, scope_type, scope_key, cost_class, target_hkd, source_file)
            VALUES (?, 'company', 'COMPANY', 'TOTAL', ?, ?)
            """,
            (2026, company_total, path.name),
        )
    insert_import_log(conn, path.name, loaded, "部门目标与公司总目标")
    return loaded


def load_project_progress(conn) -> int:
    path = SOURCE_FILES["project_progress"]
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [normalize_text(x) for x in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: i for i, name in enumerate(header)}
    loaded = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        project_code = normalize_text(row[idx["项目编码"]])
        if not project_code.startswith("R"):
            continue
        project_name = normalize_text(row[idx["项目名称"]])
        dept_name = normalize_text(row[idx["部门"]])
        category = normalize_cost_class(normalize_text(row[idx["Lv1分类"]]))
        budget = to_float(row[idx["2025年预算金额（万港币）"]])
        completed_10 = to_float(row[idx["10月实际完成数"]])
        expected_11 = to_float(row[idx["11月预计累计完成数"]])
        dept_code = ensure_department(conn, dept_name)
        upsert_project(
            conn,
            project_code=project_code,
            project_name=project_name or project_code,
            dept_code=dept_code,
            category=category,
            plan_year=2025,
            source_file=path.name,
        )
        if budget is not None:
            conn.execute(
                """
                INSERT INTO kpi_targets (year, scope_type, scope_key, cost_class, target_hkd, source_file)
                VALUES (?, 'project', ?, ?, ?, ?)
                """,
                (2025, project_code, category, budget * 10000, path.name),
            )
        if completed_10 is not None:
            conn.execute(
                """
                INSERT INTO monthly_costs
                (month, year, project_code, dept_code, cost_class, cost_component, amount_hkd, source_file, raw_tag)
                VALUES ('2025-10', 2025, ?, ?, ?, 'ROLLUP', ?, ?, 'progress_actual_10')
                """,
                (project_code, dept_code, category, completed_10 * 10000, path.name),
            )
        if expected_11 is not None:
            conn.execute(
                """
                INSERT INTO progress_forecasts
                (month, project_code, dept_code, expected_hkd, actual_hkd, note, source_file)
                VALUES ('2025-11', ?, ?, ?, NULL, ?, ?)
                """,
                (project_code, dept_code, expected_11 * 10000, "来自双周报模板预测", path.name),
            )
        loaded += 1
    insert_import_log(conn, path.name, loaded, "项目基础档案、预算、进展预测")
    return loaded


def load_labor_monthly(conn) -> int:
    path = SOURCE_FILES["labor_monthly"]
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [normalize_text(x) for x in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: i for i, name in enumerate(header)}
    month_map = {
        "人工成本-8月份": "2025-08",
        "人工成本-9月份": "2025-09",
        "人工成本-10月份": "2025-10",
        "人工成本-11月份": "2025-11",
    }
    loaded = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        project_code = normalize_text(row[idx["项目编码"]]) if idx.get("项目编码") is not None else ""
        dept_name = normalize_text(row[idx["责任部门"]]) if idx.get("责任部门") is not None else ""
        project_name = normalize_text(row[idx["项目名称"]]) if idx.get("项目名称") is not None else ""
        if not project_code.startswith("R"):
            continue
        dept_code = ensure_department(conn, dept_name)
        upsert_project(
            conn,
            project_code=project_code,
            project_name=project_name or project_code,
            dept_code=dept_code,
            category="OPEX",
            plan_year=2025,
            source_file=path.name,
        )
        for col_name, month in month_map.items():
            amount = to_float(row[idx[col_name]]) if idx.get(col_name) is not None else None
            if amount is None:
                continue
            conn.execute(
                """
                INSERT INTO monthly_costs
                (month, year, project_code, dept_code, cost_class, cost_component, amount_hkd, source_file, raw_tag)
                VALUES (?, ?, ?, ?, 'OPEX', 'LABOR', ?, ?, ?)
                """,
                (month, int(month[:4]), project_code, dept_code, amount * 10000, path.name, col_name),
            )
            loaded += 1
    insert_import_log(conn, path.name, loaded, "人工成本按月明细")
    return loaded


def load_timesheet(conn) -> int:
    path = SOURCE_FILES["timesheet"]
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [normalize_text(x) for x in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: i for i, name in enumerate(header)}
    loaded = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        report_date = row[idx["工时申报日期"]]
        project_text = normalize_text(row[idx["项目名称"]])
        employee = normalize_text(row[idx["兼职研发人员"]])
        dept_name = normalize_text(row[4])  # 源文件第5列没有标题但有部门
        month_cn = normalize_text(row[idx["月份"]])
        declared_hours = to_float(row[idx["当月研发工时数目"]])
        if not employee or declared_hours is None:
            continue
        dt_obj = report_date if isinstance(report_date, datetime) else None
        month = month_from_chinese(month_cn, fallback=dt_obj)
        project_code = extract_project_code(project_text)
        dept_code = ensure_department(conn, dept_name or "未分配部门")
        if project_code:
            upsert_project(
                conn,
                project_code=project_code,
                project_name=project_text or project_code,
                dept_code=dept_code,
                category="OPEX",
                plan_year=2026,
                source_file=path.name,
            )
        conn.execute(
            """
            INSERT INTO labor_timesheets
            (report_date, month, employee_name, project_code, project_name, dept_code, declared_hours,
             manager_approved, dept_approved, rd_approved, hr_approved, finance_approved, source_file)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                dt_obj.isoformat() if dt_obj else normalize_text(report_date),
                month,
                employee,
                project_code,
                project_text,
                dept_code,
                declared_hours,
                normalize_stage_flag(row[idx["项目负责人审批"]]),
                normalize_stage_flag(row[idx["部门主管审批"]]),
                normalize_stage_flag(row[idx["研发部门(研发管理人员)确认汇总"]]),
                normalize_stage_flag(row[idx["人力资源部完成收集工时"]]),
                0,
                path.name,
            ),
        )
        loaded += 1
    insert_import_log(conn, path.name, loaded, "工时申报与审批状态")
    return loaded


def load_dashboard_snapshot(conn) -> int:
    path = SOURCE_FILES["dashboard"]
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    loaded = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        dept_name = normalize_text(row[1])
        target = to_float(row[2])
        completed_1031 = to_float(row[3])
        completed_1113 = to_float(row[5])
        if not dept_name:
            continue
        dept_code = ensure_department(conn, dept_name)
        if target is not None:
            conn.execute(
                """
                INSERT INTO kpi_targets (year, scope_type, scope_key, cost_class, target_hkd, source_file)
                VALUES (2025, 'department', ?, 'TOTAL', ?, ?)
                """,
                (dept_code, target * 10000, path.name),
            )
        if completed_1031 is not None:
            conn.execute(
                """
                INSERT INTO monthly_costs
                (month, year, project_code, dept_code, cost_class, cost_component, amount_hkd, source_file, raw_tag)
                VALUES ('2025-10', 2025, NULL, ?, 'TOTAL', 'ROLLUP', ?, ?, 'dashboard_1031')
                """,
                (dept_code, completed_1031 * 10000, path.name),
            )
        if completed_1113 is not None:
            conn.execute(
                """
                INSERT INTO monthly_costs
                (month, year, project_code, dept_code, cost_class, cost_component, amount_hkd, source_file, raw_tag)
                VALUES ('2025-11', 2025, NULL, ?, 'TOTAL', 'ROLLUP', ?, ?, 'dashboard_1113')
                """,
                (dept_code, completed_1113 * 10000, path.name),
            )
        loaded += 1
    insert_import_log(conn, path.name, loaded, "部门看板快照")
    return loaded


def load_finance_snapshot(conn, key: str) -> int:
    path = SOURCE_FILES[key]
    month = month_from_filename(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    loaded = 0
    for row in ws.iter_rows(min_row=1, values_only=True):
        project_code = normalize_text(row[1])
        if not project_code.startswith("R"):
            continue
        project_name = normalize_text(row[2]) or project_code
        plan_year = int(to_float(row[3]) or 2025)
        category = normalize_cost_class(normalize_text(row[4]))
        outsourced_hkd = to_float(row[5])
        labor_hkd = to_float(row[6])
        other_hkd = to_float(row[7])
        total_hkd = to_float(row[8])
        total_cny = to_float(row[9])
        proj_row = conn.execute(
            "SELECT dept_code FROM projects WHERE project_code = ?",
            (project_code,),
        ).fetchone()
        dept_code = proj_row["dept_code"] if proj_row else None
        upsert_project(
            conn,
            project_code=project_code,
            project_name=project_name,
            dept_code=dept_code,
            category=category,
            plan_year=plan_year,
            source_file=path.name,
        )
        for component, amount in (
            ("OUTSOURCED", outsourced_hkd),
            ("LABOR", labor_hkd),
            ("OTHER", other_hkd),
            ("ROLLUP", total_hkd),
        ):
            if amount is None:
                continue
            conn.execute(
                """
                INSERT INTO monthly_costs
                (month, year, project_code, dept_code, cost_class, cost_component, amount_hkd, amount_cny, source_file, raw_tag)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    month,
                    int(month[:4]),
                    project_code,
                    dept_code,
                    category,
                    component,
                    amount,
                    total_cny if component == "ROLLUP" else None,
                    path.name,
                    key,
                ),
            )
            loaded += 1
    insert_import_log(conn, path.name, loaded, f"财务快照 {month}")
    return loaded


def build_company_rollup(conn) -> None:
    rows = conn.execute(
        """
        SELECT month, year, cost_class, cost_component, SUM(amount_hkd) AS total
        FROM monthly_costs
        GROUP BY month, year, cost_class, cost_component
        """
    ).fetchall()
    for row in rows:
        conn.execute(
            """
            INSERT INTO monthly_costs
            (month, year, project_code, dept_code, cost_class, cost_component, amount_hkd, source_file, raw_tag)
            VALUES (?, ?, NULL, NULL, ?, ?, ?, 'SYSTEM', 'company_rollup')
            """,
            (row["month"], row["year"], row["cost_class"], row["cost_component"], row["total"]),
        )


def load_all_sources() -> dict[str, int]:
    with db_session() as conn:
        reset_business_tables(conn)
        stats = {
            "dept_target": load_dept_target(conn),
            "project_progress": load_project_progress(conn),
            "labor_monthly": load_labor_monthly(conn),
            "timesheet": load_timesheet(conn),
            "dashboard": load_dashboard_snapshot(conn),
            "finance_1031": load_finance_snapshot(conn, "finance_1031"),
            "finance_prev": load_finance_snapshot(conn, "finance_prev"),
        }
        build_company_rollup(conn)
        return stats
